import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from PIL import Image as PilImage
import os

class EquipmentInspectionApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Inspeção de Equipamentos")
        self.geometry("800x600")
        self.style = ttk.Style(self)
        self.style.theme_use("clam")

        # Cria e organiza os widgets principais
        self.create_widgets()

    def create_widgets(self):
        # Cria um notebook com abas
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(expand=True, fill=tk.BOTH)

        # Abas do notebook
        self.tab_imagem = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_imagem, text='Seleção de Imagem')

        self.tab_trabalho = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_trabalho, text='Trabalho Realizado')

        # Cria widgets para cada aba
        self.create_imagem_widgets()
        self.create_trabalho_widgets()

    def create_imagem_widgets(self):
        """
        Cria widgets para a aba de seleção de imagem.
        """
        # Label para instruções
        label_imagem = ttk.Label(self.tab_imagem, text="Selecione uma ou mais imagens:", font=('Arial', 12))
        label_imagem.pack(pady=20)

        # Variável para armazenar os caminhos das imagens selecionadas
        self.selected_images = []

        def select_images():
            """
            Abre um diálogo para selecionar múltiplas imagens e atualiza a visualização.
            """
            filenames = filedialog.askopenfilenames(filetypes=[("Imagens", "*.png;*.jpg;*.jpeg")])
            if filenames:
                self.selected_images = filenames
                label_preview.config(text=f"Imagens selecionadas: {', '.join(filenames)}")

        # Botão para selecionar imagens
        btn_select_images = ttk.Button(self.tab_imagem, text="Selecionar Imagens", command=select_images)
        btn_select_images.pack(pady=10)

        # Label para mostrar os caminhos das imagens selecionadas
        label_preview = ttk.Label(self.tab_imagem, text="", font=('Arial', 10))
        label_preview.pack()

        # Botão para ir para a próxima aba
        btn_next = ttk.Button(self.tab_imagem, text="Próximo", command=lambda: self.notebook.select(self.tab_trabalho))
        btn_next.pack(pady=20)

    def create_trabalho_widgets(self):
        """
        Cria widgets para a aba de trabalho realizado.
        """
        fields_trabalho = [
            "CLIENTE", "TIPO DE EQUIPAMENTO", "Equipamentos/Componentes", "NOME DO INSPETOR"
        ]

        # Dicionário para armazenar as variáveis de entrada
        self.fields_entries_trabalho = {}

        for idx, field in enumerate(fields_trabalho):
            # Label para cada campo
            label = ttk.Label(self.tab_trabalho, text=field + ": ", font=('Arial', 12))
            label.grid(row=idx, column=0, padx=10, pady=5, sticky=tk.E)

            if field in ["CLIENTE", "TIPO DE EQUIPAMENTO", "Equipamentos/Componentes"]:
                # Cria combobox para campos com opções pré-definidas
                options = {
                    "CLIENTE": ["BRK_Q3", "ACELEN_BA", "EQUINOR - WHA"],
                    "TIPO DE EQUIPAMENTO": ["Circuito", "Trocador", "Compressor"],
                    "Equipamentos/Componentes": ["IA-6-H2X", "CASCO", "EIXO DO COMPRESSOR", "O-4-18-13011 (A1A1)", "CASCO", "6-PT-31214_SHT01", "6-PT-31224_SHT05", "4-PT-31225_SHT04"]
                }[field]
                var = tk.StringVar()
                entry = ttk.Combobox(self.tab_trabalho, width=40, textvariable=var, values=options, state='readonly')
                entry.current(0)
                entry.grid(row=idx, column=1, padx=10, pady=5, sticky=tk.W)
                self.fields_entries_trabalho[field] = var

            elif field == "NOME DO INSPETOR":
                # Cria combobox para seleção do nome do inspetor
                options = ["João 1", "João 2", "João 3", "Bruno 1", "Bruno 2", "Bruno 3", "Jessé"]
                var = tk.StringVar()
                var.set(options[0])
                entry = ttk.Combobox(self.tab_trabalho, width=40, textvariable=var, values=options, state='readonly')
                entry.grid(row=idx, column=1, padx=10, pady=5, sticky=tk.W)
                self.fields_entries_trabalho[field] = var

        # Botão para exportar dados para Excel
        self.btn_export = ttk.Button(self.tab_trabalho, text="Exportar para Excel", command=self.export_to_excel)
        self.btn_export.grid(row=len(fields_trabalho), column=0, columnspan=2, pady=10)

    def export_to_excel(self):
        """
        Exporta os dados preenchidos para um arquivo Excel.
        """
        # Cria um novo workbook e uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Inspeção de Equipamentos"

        # Estilos para a planilha
        header_font = Font(bold=True, size=14)
        cell_font = Font(size=12)
        center_alignment = Alignment(horizontal='center', vertical='center')
        border_style = Border(left=Side(border_style="thin"), 
                               right=Side(border_style="thin"), 
                               top=Side(border_style="thin"), 
                               bottom=Side(border_style="thin"))

        # Adiciona o título da planilha
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = 'Inspeção de Equipamentos'
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = center_alignment

        # Adiciona uma linha vazia para espaçamento
        ws.append([])

        # Cabeçalhos das colunas
        headers = ['Campo', 'Valor']
        ws.append(headers)
        for cell in ws[2]:
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border_style

        # Adiciona os dados dos campos
        for field, var in self.fields_entries_trabalho.items():
            value = var.get()
            ws.append([field, value])
            # Aplique o estilo às células da linha adicionada
            for cell in ws[ws.max_row]:
                cell.font = cell_font
                cell.alignment = center_alignment
                cell.border = border_style

        # Adiciona o campo de "Registro fotográfico"
        row_start = ws.max_row + 2  # Linha onde começará o campo de "Registro fotográfico"
        ws[f'A{row_start}'] = 'Registro fotográfico'
        ws[f'A{row_start}'].font = Font(bold=True, size=14)
        ws[f'A{row_start}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{row_start}'].border = border_style

        # Adiciona as imagens abaixo do campo de "Registro fotográfico"
        row_start += 1  # Linha onde começará a adicionar imagens
        if self.selected_images:
            for image_path in self.selected_images:
                try:
                    # Adiciona a imagem à planilha
                    excel_img = Image(image_path)
                    excel_img.anchor = f'A{row_start}'
                    ws.add_image(excel_img)
                    
                    # Ajusta a linha para a próxima imagem
                    row_start += 20  # Ajusta a linha com base no tamanho das imagens

                except Exception as e:
                    messagebox.showerror("Erro", f"Não foi possível adicionar a imagem {image_path}: {e}")

        # Ajusta o tamanho das colunas com base no conteúdo
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Salva o arquivo Excel
        try:
            wb.save('dados_inspecao.xlsx')
            messagebox.showinfo("Exportação Concluída", "Dados exportados para dados_inspecao.xlsx")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar o arquivo Excel: {e}")

if __name__ == "__main__":
    app = EquipmentInspectionApp()
    app.mainloop()
