import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from PIL import Image as PilImage
import os

class EquipmentInspectionApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Inspeção de Equipamentos")
        self.geometry("800x600")
        self.style = ttk.Style(self)
        self.style.theme_use("clam")

        # Carrega os dados do cliente
        self.clientes_data = self.load_clientes_data()

        # Carrega os dados do tipo de equipamento
        self.tipo_equipamento_options = self.load_tipo_equipamento_data()

        # Carrega os dados do componente
        self.componente_data = self.load_componente_data()

        # Cria e organiza os widgets principais
        self.create_widgets()

    def load_clientes_data(self):
        clientes = {}
        try:
            wb = load_workbook('BancoDeDadosAraujo.xlsx')
            ws = wb['Clientes']
            for row in ws.iter_rows(min_row=2, values_only=True):  # Assumindo que a primeira linha é o cabeçalho
                clientes[row[1]] = row[0]  # Assumindo que o nome do cliente está na segunda coluna e o código na primeira coluna
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível carregar os dados dos clientes: {e}")
        return clientes
    
    def load_tipo_equipamento_data(self):
        tipoequipamento = []
        try:
            wb = load_workbook('BancoDeDadosAraujo.xlsx')
            ws = wb['TabTipoEquipamento']
            for row in ws.iter_rows(min_row=2, values_only=True):  # Assumindo que a primeira linha é o cabeçalho
                tipoequipamento.append(row[1])  # Assume que o tipo de equipamento está na segunda coluna
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível carregar os dados do tipo de equipamento: {e}")
        return tipoequipamento
    
    def load_componente_data(self):
        componentes = []
        try:
            wb = load_workbook('BancoDeDadosAraujo.xlsx')
            ws = wb['EquipamentosComponentes']
            for row in ws.iter_rows(min_row=2, values_only=True):  # Assumindo que a primeira linha é o cabeçalho
                componentes.append((row[3], row[0]))  # Assume que o nome do componente está na quarta coluna e o código do cliente na primeira coluna
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível carregar os dados dos componentes: {e}")
        return componentes

    def create_widgets(self):
        # Cria um notebook com abas
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(expand=True, fill=tk.BOTH)

        # Abas do notebook
        self.tab_imagem = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_imagem, text='Seleção de Imagem')

        self.tab_trabalho = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_trabalho, text='Trabalho Realizado')

        # Cria widgets para cada aba
        self.create_imagem_widgets()
        self.create_trabalho_widgets()

    def create_imagem_widgets(self):
        """
        Cria widgets para a aba de seleção de imagem.
        """
        # Label para instruções
        label_imagem = ttk.Label(self.tab_imagem, text="Selecione uma ou mais imagens:", font=('Arial', 12))
        label_imagem.pack(pady=20)

        # Variável para armazenar os caminhos das imagens selecionadas
        self.selected_images = []

        def select_images():
            """
            Abre um diálogo para selecionar múltiplas imagens e atualiza a visualização.
            """
            filenames = filedialog.askopenfilenames(filetypes=[("Imagens", "*.png;*.jpg;*.jpeg")])
            if filenames:
                self.selected_images = filenames
                label_preview.config(text=f"Imagens selecionadas: {', '.join(filenames)}")

        # Botão para selecionar imagens
        btn_select_images = ttk.Button(self.tab_imagem, text="Selecionar Imagens", command=select_images)
        btn_select_images.pack(pady=10)

        # Label para mostrar os caminhos das imagens selecionadas
        label_preview = ttk.Label(self.tab_imagem, text="", font=('Arial', 10))
        label_preview.pack()

        # Botão para ir para a próxima aba
        btn_next = ttk.Button(self.tab_imagem, text="Próximo", command=lambda: self.notebook.select(self.tab_trabalho))
        btn_next.pack(pady=20)

    def create_trabalho_widgets(self):
        """
        Cria widgets para a aba de trabalho realizado.
        """
        fields_trabalho = [
            "CLIENTE", "TIPO DE EQUIPAMENTO", "COMPONENTES", "NOME DO INSPETOR"
        ]

        # Dicionário para armazenar as variáveis de entrada
        self.fields_entries_trabalho = {}

        for idx, field in enumerate(fields_trabalho):
            # Label para cada campo
            label = ttk.Label(self.tab_trabalho, text=field + ": ", font=('Arial', 12))
            label.grid(row=idx, column=0, padx=10, pady=5, sticky=tk.E)

            if field == "CLIENTE":
                # Cria combobox para seleção do cliente
                var = tk.StringVar()
                entry = ttk.Combobox(self.tab_trabalho, width=40, textvariable=var, values=list(self.clientes_data.keys()), state='readonly')
                entry.current(0)
                entry.grid(row=idx, column=1, padx=10, pady=5, sticky=tk.W)
                self.fields_entries_trabalho[field] = var

                # Adiciona evento para atualizar componentes quando o cliente é selecionado
                entry.bind("<<ComboboxSelected>>", self.update_componentes)

            elif field == "TIPO DE EQUIPAMENTO":
                # Cria combobox para seleção do tipo de equipamento
                var = tk.StringVar()
                entry = ttk.Combobox(self.tab_trabalho, width=40, textvariable=var, values=self.tipo_equipamento_options, state='readonly')
                entry.current(0)
                entry.grid(row=idx, column=1, padx=10, pady=5, sticky=tk.W)
                self.fields_entries_trabalho[field] = var

            elif field == "COMPONENTES":
                # Cria combobox para seleção do componente
                var = tk.StringVar()
                self.entry_componentes = ttk.Combobox(self.tab_trabalho, width=40, textvariable=var, state='readonly')
                self.entry_componentes.grid(row=idx, column=1, padx=10, pady=5, sticky=tk.W)
                self.fields_entries_trabalho[field] = var

            elif field == "NOME DO INSPETOR":
                # Cria combobox para seleção do nome do inspetor
                options = ["João 1", "João 2", "João 3", "Bruno 1", "Bruno 2", "Bruno 3", "Jessé"]
                var = tk.StringVar()
                var.set(options[0])
                entry = ttk.Combobox(self.tab_trabalho, width=40, textvariable=var, values=options, state='readonly')
                entry.grid(row=idx, column=1, padx=10, pady=5, sticky=tk.W)
                self.fields_entries_trabalho[field] = var

        # Botão para exportar dados para Excel
        self.btn_export = ttk.Button(self.tab_trabalho, text="Exportar para Excel", command=self.export_to_excel)
        self.btn_export.grid(row=len(fields_trabalho), column=0, columnspan=2, pady=10)

    def update_componentes(self, event):
        """
        Atualiza a lista de componentes disponíveis com base no cliente selecionado.
        """
        cliente_nome = self.fields_entries_trabalho["CLIENTE"].get()
        cliente_codigo = self.clientes_data.get(cliente_nome)
        
        if cliente_codigo:
            componentes_filtrados = [comp[0] for comp in self.componente_data if comp[1] == cliente_codigo]
            self.entry_componentes['values'] = componentes_filtrados
            self.entry_componentes.current(0)

    def export_to_excel(self):
        """
        Exporta os dados preenchidos para um arquivo Excel.
        """
        # Cria um novo workbook e uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Inspeção de Equipamentos"

        # Estilos para a planilha
        header_font = Font(bold=True, size=14)
        cell_font = Font(size=12)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # Cabeçalho
        for idx, field in enumerate(self.fields_entries_trabalho.keys()):
            cell = ws.cell(row=1, column=idx+1, value=field)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Dados
        for idx, (field, var) in enumerate(self.fields_entries_trabalho.items()):
            cell = ws.cell(row=2, column=idx+1, value=var.get())
            cell.font = cell_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Salva o workbook
        filename = "Inspecao_Equipamentos.xlsx"
        try:
            wb.save(filename)
            messagebox.showinfo("Sucesso", f"Dados exportados com sucesso para {filename}")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar o arquivo: {e}")

if __name__ == "__main__":
    app = EquipmentInspectionApp()
    app.mainloop()
